import openpyxl, os

workbook = openpyxl.load_workbook(os.getcwd() + "/AutomateTasks/example.xlsx")
print(workbook.get_sheet_names())
sheet = workbook["Sheet1"]
cell = sheet["A1"]
print(cell.value)
print(type(cell.value))
cell = sheet.cell(1,2)
print(cell.value)
for i in range(1,8):
    print(i, sheet.cell(i,2).value)